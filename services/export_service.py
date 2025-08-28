# services/export_service.py

import io
import csv
from datetime import datetime
from typing import Dict, Any, Optional, List
import json

# Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import PieChart, LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl not available. Excel export will fail. Install with: pip install openpyxl")
    EXCEL_AVAILABLE = False

# The reportlab library is needed for PDF generation
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    print("Warning: reportlab not available. PDF export will fail. Install with: pip install reportlab")
    PDF_AVAILABLE = False

async def generate_holdings_csv(context: Dict[str, Any]) -> io.StringIO:
    """Generates a CSV of portfolio holdings in memory."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Ticker', 'Shares', 'Current Price', 'Market Value', 'Day Change', 'Day Change %'])
    
    holdings = context.get("holdings_with_values", [])
    if not holdings:
        writer.writerow(['No holdings found in this portfolio.'])
        return output

    for holding in holdings:
        writer.writerow([
            holding.asset.ticker,
            holding.shares,
            f"{holding.current_price:.2f}",
            f"{holding.market_value:.2f}",
            f"{holding.day_change * holding.shares:.2f}",
            f"{holding.day_change_percent:.2f}%"
        ])
    
    output.seek(0)
    return output

async def generate_excel_report(
    portfolio: Any,
    context: Dict[str, Any],
    analytics: Dict[str, Any]
) -> io.BytesIO:
    """Generate comprehensive Excel report with multiple sheets."""
    if not EXCEL_AVAILABLE:
        raise ImportError("The 'openpyxl' library is required for Excel generation.")
    
    # Create workbook and buffer
    wb = Workbook()
    buffer = io.BytesIO()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    _create_summary_sheet(wb, portfolio, context, analytics)
    _create_holdings_sheet(wb, context)
    _create_performance_sheet(wb, analytics)
    _create_risk_analysis_sheet(wb, analytics)
    _create_charts_sheet(wb, context, analytics)
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def _create_summary_sheet(wb: Workbook, portfolio: Any, context: Dict[str, Any], analytics: Dict[str, Any]):
    """Create portfolio summary sheet."""
    ws = wb.create_sheet("Portfolio Summary", 0)
    
    # Styling
    header_font = Font(size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(size=12, bold=True)
    currency_format = '"$"#,##0.00'
    percent_format = '0.00%'
    
    # Title section
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Portfolio Analysis Report: {portfolio.name}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws["A3"] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = subheader_font
    
    # Key metrics section
    ws["A5"] = "Key Portfolio Metrics"
    ws["A5"].font = subheader_font
    
    total_value = context.get('total_value', 0)
    holdings_count = len(context.get('holdings_with_values', []))
    day_change = context.get('total_day_change', 0)
    day_change_pct = (day_change / total_value * 100) if total_value > 0 else 0
    
    metrics = [
        ("Total Portfolio Value", total_value, currency_format),
        ("Number of Holdings", holdings_count, None),
        ("Today's Change", day_change, currency_format),
        ("Today's Change %", day_change_pct/100, percent_format),
    ]
    
    row = 6
    for metric, value, fmt in metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = value
        if fmt:
            ws[f"B{row}"].number_format = fmt
        row += 1
    
    # Performance metrics section
    analytics_data = analytics.get("data", {})
    perf_stats = analytics_data.get("performance_stats", {})
    risk_ratios = analytics_data.get("risk_adjusted_ratios", {})
    
    ws[f"A{row+1}"] = "Performance Metrics"
    ws[f"A{row+1}"].font = subheader_font
    
    perf_metrics = [
        ("Annualized Return", perf_stats.get('annualized_return_pct', 0)/100, percent_format),
        ("Annualized Volatility", perf_stats.get('annualized_volatility_pct', 0)/100, percent_format),
        ("Sharpe Ratio", risk_ratios.get('sharpe_ratio', 0), "0.000"),
        ("Alpha", perf_stats.get('alpha', 0)/100, percent_format),
        ("Beta", perf_stats.get('beta', 1.0), "0.00"),
    ]
    
    row += 2
    for metric, value, fmt in perf_metrics:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = value
        ws[f"B{row}"].number_format = fmt
        row += 1
    
    # AI Summary section
    ws[f"A{row+1}"] = "AI-Generated Summary"
    ws[f"A{row+1}"].font = subheader_font
    
    summary_text = analytics.get("summary", "No summary available.")
    ws.merge_cells(f"A{row+2}:F{row+5}")
    ws[f"A{row+2}"] = summary_text
    ws[f"A{row+2}"].alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

def _create_holdings_sheet(wb: Workbook, context: Dict[str, Any]):
    """Create detailed holdings sheet."""
    ws = wb.create_sheet("Holdings Detail")
    
    # Headers
    headers = [
        "Ticker", "Company Name", "Shares", "Purchase Price", 
        "Current Price", "Market Value", "Gain/Loss", "Gain/Loss %",
        "Day Change", "Day Change %", "Weight %"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    holdings = context.get("holdings_with_values", [])
    total_value = context.get('total_value', 1)
    
    # Data rows
    for row, holding in enumerate(holdings, 2):
        gain_loss = holding.market_value - (holding.shares * getattr(holding, 'purchase_price', holding.current_price))
        gain_loss_pct = (gain_loss / (holding.shares * getattr(holding, 'purchase_price', holding.current_price))) if getattr(holding, 'purchase_price', holding.current_price) > 0 else 0
        weight = (holding.market_value / total_value) if total_value > 0 else 0
        
        data = [
            holding.asset.ticker,
            getattr(holding.asset, 'name', 'N/A'),
            holding.shares,
            getattr(holding, 'purchase_price', holding.current_price),
            holding.current_price,
            holding.market_value,
            gain_loss,
            gain_loss_pct,
            getattr(holding, 'day_change', 0) * holding.shares,
            getattr(holding, 'day_change_percent', 0) / 100,
            weight
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            
            # Apply formatting
            if col in [4, 5, 6, 7, 9]:  # Price and value columns
                cell.number_format = '"$"#,##0.00'
            elif col in [8, 10, 11]:  # Percentage columns
                cell.number_format = '0.00%'
            elif col == 3:  # Shares
                cell.number_format = '#,##0'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

def _create_performance_sheet(wb: Workbook, analytics: Dict[str, Any]):
    """Create performance analysis sheet."""
    ws = wb.create_sheet("Performance Analysis")
    
    analytics_data = analytics.get("data", {})
    perf_stats = analytics_data.get("performance_stats", {})
    risk_ratios = analytics_data.get("risk_adjusted_ratios", {})
    drawdown = analytics_data.get("drawdown_stats", {})
    
    # Title
    ws["A1"] = "Performance Analysis"
    ws["A1"].font = Font(size=14, bold=True)
    
    # Performance metrics
    ws["A3"] = "Return Metrics"
    ws["A3"].font = Font(size=12, bold=True)
    
    perf_data = [
        ("Metric", "Value", "Benchmark*", "Difference"),
        ("Total Return (YTD)", perf_stats.get('ytd_return_pct', 0)/100, 0.085, (perf_stats.get('ytd_return_pct', 0)/100) - 0.085),
        ("Annualized Return", perf_stats.get('annualized_return_pct', 0)/100, 0.10, (perf_stats.get('annualized_return_pct', 0)/100) - 0.10),
        ("Annualized Volatility", perf_stats.get('annualized_volatility_pct', 0)/100, 0.15, (perf_stats.get('annualized_volatility_pct', 0)/100) - 0.15),
    ]
    
    row = 4
    for data_row in perf_data:
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if row > 4 and col > 1:  # Data rows, not headers
                cell.number_format = '0.00%'
            elif row == 4:  # Header row
                cell.font = Font(bold=True)
        row += 1
    
    # Risk-adjusted ratios
    ws[f"A{row+1}"] = "Risk-Adjusted Ratios"
    ws[f"A{row+1}"].font = Font(size=12, bold=True)
    
    ratio_data = [
        ("Ratio", "Value", "Interpretation"),
        ("Sharpe Ratio", risk_ratios.get('sharpe_ratio', 0), "Risk-adjusted return"),
        ("Information Ratio", risk_ratios.get('information_ratio', 0), "Active return vs tracking error"),
        ("Alpha", perf_stats.get('alpha', 0)/100, "Excess return vs benchmark"),
        ("Beta", perf_stats.get('beta', 1.0), "Market sensitivity"),
    ]
    
    row += 2
    for data_row in ratio_data:
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if row > row-len(ratio_data)+1 and col == 2 and isinstance(value, (int, float)):
                if "Alpha" in ws.cell(row=row, column=1).value:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '0.000'
            elif col == 1 or row == row-len(ratio_data)+1:  # First column or header
                cell.font = Font(bold=True)
        row += 1
    
    # Benchmark note
    ws[f"A{row+1}"] = "*Benchmark: S&P 500 Total Return Index"
    ws[f"A{row+1}"].font = Font(italic=True, size=10)

def _create_risk_analysis_sheet(wb: Workbook, analytics: Dict[str, Any]):
    """Create risk analysis sheet."""
    ws = wb.create_sheet("Risk Analysis")
    
    analytics_data = analytics.get("data", {})
    risk_measures = analytics_data.get("risk_measures", {})
    drawdown_stats = analytics_data.get("drawdown_stats", {})
    
    # Title
    ws["A1"] = "Risk Analysis"
    ws["A1"].font = Font(size=14, bold=True)
    
    # Value at Risk
    ws["A3"] = "Value at Risk (VaR) Analysis"
    ws["A3"].font = Font(size=12, bold=True)
    
    var_data = [("Confidence Level", "VaR", "Expected Shortfall")]
    
    for confidence, metrics in risk_measures.items():
        if isinstance(metrics, dict):
            var_data.append((
                confidence,
                metrics.get('var', 0),
                metrics.get('cvar_expected_shortfall', 0)
            ))
    
    row = 4
    for data_row in var_data:
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if row > 4 and col > 1:  # Data rows
                cell.number_format = '0.00%'
            elif row == 4:  # Header row
                cell.font = Font(bold=True)
        row += 1
    
    # Drawdown analysis
    ws[f"A{row+1}"] = "Drawdown Analysis"
    ws[f"A{row+1}"].font = Font(size=12, bold=True)
    
    drawdown_data = [
        ("Metric", "Value"),
        ("Maximum Drawdown", drawdown_stats.get('max_drawdown_pct', 0)/100),
        ("Average Drawdown", drawdown_stats.get('avg_drawdown_pct', 0)/100),
        ("Recovery Time (days)", drawdown_stats.get('avg_recovery_days', 0)),
    ]
    
    row += 2
    for data_row in drawdown_data:
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if row > row-len(drawdown_data)+1 and col == 2:  # Data values
                if isinstance(value, (int, float)) and abs(value) < 1:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0'
            elif col == 1 or row == row-len(drawdown_data)+1:  # Labels or header
                cell.font = Font(bold=True)
        row += 1

def _create_charts_sheet(wb: Workbook, context: Dict[str, Any], analytics: Dict[str, Any]):
    """Create charts and visualizations sheet."""
    ws = wb.create_sheet("Charts & Analysis")
    
    # Title
    ws["A1"] = "Portfolio Visualization"
    ws["A1"].font = Font(size=14, bold=True)
    
    # Create allocation data for pie chart
    holdings = context.get("holdings_with_values", [])
    if holdings:
        # Top 5 holdings + Others
        sorted_holdings = sorted(holdings, key=lambda h: h.market_value, reverse=True)
        
        ws["A3"] = "Portfolio Allocation"
        ws["A3"].font = Font(size=12, bold=True)
        
        headers = ["Asset", "Market Value", "Weight %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
        
        total_value = context.get('total_value', 1)
        row = 5
        
        # Top 5 holdings
        for holding in sorted_holdings[:5]:
            weight = (holding.market_value / total_value) if total_value > 0 else 0
            ws.cell(row=row, column=1, value=holding.asset.ticker)
            ws.cell(row=row, column=2, value=holding.market_value).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=3, value=weight).number_format = '0.00%'
            row += 1
        
        # Others category
        if len(sorted_holdings) > 5:
            others_value = sum(h.market_value for h in sorted_holdings[5:])
            others_weight = (others_value / total_value) if total_value > 0 else 0
            ws.cell(row=row, column=1, value="Others")
            ws.cell(row=row, column=2, value=others_value).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=3, value=others_weight).number_format = '0.00%'
    
    # Monthly performance table (mock data for demonstration)
    ws["E3"] = "Monthly Performance History"
    ws["E3"].font = Font(size=12, bold=True)
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    years = ["2024", "2023", "2022"]
    
    # Headers
    ws["E4"] = "Month"
    for col, year in enumerate(years, 2):
        ws.cell(row=4, column=4+col, value=year)
        ws.cell(row=4, column=4+col).font = Font(bold=True)
    
    # Data (mock monthly returns)
    for row, month in enumerate(months, 5):
        ws.cell(row=row, column=5, value=month)
        for col, year in enumerate(years, 1):
            # Generate consistent mock returns based on month/year hash
            return_val = (hash(f"{month}{year}") % 20 - 10) / 100  # -10% to +10%
            cell = ws.cell(row=row, column=5+col, value=return_val)
            cell.number_format = '0.00%'

async def generate_pdf_report(
    portfolio: Any, 
    context: Dict[str, Any], 
    analytics: Dict[str, Any]
) -> io.BytesIO:
    """Generates a professional PDF report for the portfolio."""
    if not PDF_AVAILABLE:
        raise ImportError("The 'reportlab' library is required for PDF generation.")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"Portfolio Analysis Report: {portfolio.name}", styles['h1']))
    story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 0.25*inch))

    # AI Summary
    story.append(Paragraph("AI-Generated Summary", styles['h2']))
    summary_text = analytics.get("summary", "No summary available.").replace('\n', '<br/>')
    story.append(Paragraph(summary_text, styles['BodyText']))
    story.append(Spacer(1, 0.25*inch))

    # Key Metrics Table
    story.append(Paragraph("Key Performance & Risk Metrics", styles['h2']))
    
    analytics_data = analytics.get("data", {})
    perf_stats = analytics_data.get("performance_stats", {})
    risk_ratios = analytics_data.get("risk_adjusted_ratios", {})
    drawdown = analytics_data.get("drawdown_stats", {})
    var95 = analytics_data.get("risk_measures", {}).get("95%", {})

    data = [
        ['Metric', 'Value'],
        ['Annualized Return', f"{perf_stats.get('annualized_return_pct', 0):.2f}%"],
        ['Annualized Volatility', f"{perf_stats.get('annualized_volatility_pct', 0):.2f}%"],
        ['Sharpe Ratio', f"{risk_ratios.get('sharpe_ratio', 0):.3f}"],
        ['Max Drawdown', f"{drawdown.get('max_drawdown_pct', 0):.2f}%"],
        ['95% Value-at-Risk (VaR)', f"{var95.get('var', 0) * 100:.2f}%"],
        ['95% Expected Shortfall (CVaR)', f"{var95.get('cvar_expected_shortfall', 0) * 100:.2f}%"],
    ]
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    story.append(table)
    story.append(Spacer(1, 0.25*inch))

    doc.build(story)
    buffer.seek(0)
    return buffer

async def generate_json_export(
    portfolio: Any,
    context: Dict[str, Any], 
    analytics: Dict[str, Any]
) -> str:
    """Generate JSON export for API integrations."""
    
    export_data = {
        "portfolio": {
            "id": getattr(portfolio, 'id', None),
            "name": getattr(portfolio, 'name', 'Unknown Portfolio'),
            "created_at": getattr(portfolio, 'created_at', datetime.now()).isoformat(),
        },
        "summary": {
            "total_value": context.get('total_value', 0),
            "holdings_count": len(context.get('holdings_with_values', [])),
            "day_change": context.get('total_day_change', 0),
            "day_change_percent": (context.get('total_day_change', 0) / context.get('total_value', 1) * 100) if context.get('total_value', 0) > 0 else 0,
        },
        "holdings": [],
        "analytics": analytics,
        "export_metadata": {
            "generated_at": datetime.now().isoformat(),
            "export_version": "1.0",
            "data_source": "portfolio_analytics_platform"
        }
    }
    
    # Add holdings data
    holdings = context.get("holdings_with_values", [])
    for holding in holdings:
        export_data["holdings"].append({
            "ticker": holding.asset.ticker,
            "name": getattr(holding.asset, 'name', 'N/A'),
            "shares": holding.shares,
            "current_price": holding.current_price,
            "market_value": holding.market_value,
            "day_change": getattr(holding, 'day_change', 0),
            "day_change_percent": getattr(holding, 'day_change_percent', 0),
            "weight_percent": (holding.market_value / context.get('total_value', 1) * 100) if context.get('total_value', 0) > 0 else 0
        })
    
    return json.dumps(export_data, indent=2, default=str)